import matplotlib
import matplotlib.pyplot as pl
matplotlib.use("TKAgg")
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure



import openpyxl
from openpyxl import *

import os
import tkinter as tki
from tkinter import *
from tkinter.font import Font
from tkinter import messagebox

import mysql.connector
import PIL
from PIL import Image,ImageTk

#######################
    



#########################
def detin():
    # opening the existing excel file 
    wb = load_workbook('details/excel.xlsx') 
  
    # create the sheet object 
    sheet = wb.active 
  
  
    def excel():
        # resize the width of columns in 
        # excel spreadsheet 
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10
        sheet.column_dimensions['O'].width = 10
        sheet.column_dimensions['P'].width = 30
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 20

        # write given data to an excel spreadsheet 
        # at particular location 
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "subject"
        sheet.cell(row=1, column=3).value = "Semester"
        sheet.cell(row=1, column=4).value = "Admn Number"
        sheet.cell(row=1, column=5).value = "Contact Number"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Address"
        sheet.cell(row=1, column=8).value = "Fathers Name"
        sheet.cell(row=1, column=9).value = "Mothers Name"
        sheet.cell(row=1, column=10).value = "Gender"
        sheet.cell(row=1, column=11).value = "Date Of Birth"
        sheet.cell(row=1, column=12).value = "Admn Category"
        sheet.cell(row=1, column=13).value = "Category"
        sheet.cell(row=1, column=14).value = "Minority"
        sheet.cell(row=1, column=15).value = "Mobile No"
        sheet.cell(row=1, column=16).value = "Email ID"
        sheet.cell(row=1, column=17).value = "Blood Group"
        sheet.cell(row=1, column=18).value = "Aadhar Card No"
 
    # Function to set focus (cursor) 
    def focus1(event):
            # set focus on the course_field box 
            Name_field.focus_set() 

    def focus2(event):
            # set focus on the course_field box 
            Subject_field.focus_set()
  
    # Function to set focus 
    def focus3(event): 
            # set focus on the sem_field box 
            sem_field.focus_set() 
  
  
    # Function to set focus 
    def focus4(event): 
             # set focus on the form_no_field box 
             Admn_No_field.focus_set() 
  
  
    # Function to set focus 
    def focus5(event): 
            # set focus on the contact_no_field box 
            contact_no_field.focus_set() 
  
  
    # Function to set focus 
    def focus6(event): 
            # set focus on the email_id_field box 
            email_id_field.focus_set() 
  
  
    # Function to set focus 
    def focus7(event): 
            # set focus on the address_field box 
            address_field.focus_set()

    def focus8(event): 
            # set focus on the form_no_field box 
            Fathers_Name_field.focus_set() 
    
    # Function to set focus 
    def focus9(event): 
            # set focus on the contact_no_field box 
            Mothers_Name_field.focus_set() 
  
  
    # Function to set focus 
    def focus10(event): 
            # set focus on the email_id_field box 
            Gender_field.focus_set() 
  
  
    # Function to set focus 
    def focus11(event): 
            # set focus on the address_field box 
            Date_Of_Birth_field.focus_set()

    # Function to set focus 
    def focus12(event): 
            # set focus on the address_field box 
            Admn_Category_field.focus_set()

    # Function to set focus 
    def focus13(event): 
            # set focus on the address_field box 
            Category_field.focus_set()

    # Function to set focus 
    def focus14(event): 
            # set focus on the address_field box 
            Minority_field.focus_set()

    def focus15(event): 
            # set focus on the address_field box 
            Mobile_No_field.focus_set()

    def focus16(event): 
            # set focus on the address_field box 
            Email_ID_field.focus_set()

    def focus17(event): 
            # set focus on the address_field box 
            Blood_Group_field.focus_set()

    def focus18(event): 
            # set focus on the address_field box 
            Aadhar_Card_No_field.focus_set()
 
  
  
        # Function for clearing the 
        # contents of text entry boxes 
    def clear():
             # clear the content of text entry box 
             Name_field.delete(0, END) 
             Subject_field.delete(0, END) 
             sem_field.delete(0, END) 
             Admn_No_field.delete(0, END) 
             contact_no_field.delete(0, END) 
             email_id_field.delete(0, END) 
             address_field.delete(0, END) 
             Fathers_Name_field.delete(0, END)
             Mothers_Name_field.delete(0, END)
             Gender_field.delete(0, END)
             Date_Of_Birth_field.delete(0, END)
             Admn_Category_field.delete(0, END)
             Category_field.delete(0, END)
             Minority_field.delete(0, END)
             Mobile_No_field.delete(0, END)
             Email_ID_field.delete(0, END)
             Blood_Group_field.delete(0, END)
             Aadhar_Card_No_field.delete(0, END)

    
        # Function to take data from GUI  
        # window and write to an excel file 
    def insert():
            # if user not fill any entry 
            # then print "empty input" 
            if (Name_field.get() == "" and
                Subject_field.get() == "" and
                sem_field.get() == "" and
                Admn_no_field.get() == "" and
                contact_no_field.get() == "" and
                email_id_field.get() == "" and
                address_field.get() == "" and 
                Fathers_Name_field.get() == "" and
                Mothers_Name_field.get() == "" and
                Gender_field.get() == "" and
                Date_Of_Birth_field.get() == "" and
                Admn_Category_field.get() == "" and
                Category_field.get() == "" and
                Minority_field.get() == "" and
                Mobile_No_field.get() == "" and
                Email_ID_field.get() == "" and
                Blood_Group_field.get() == "" and
                Aadhar_Card_No_field.get() == ""):

                print("empty input") 
  
            else:
                # assigning the max row and max column 
                # value upto which data is written 
                # in an excel sheet to the variable 
                current_row = sheet.max_row 
                current_column = sheet.max_column 
  
                # get method returns current text 
                # as string which we write into 
                # excel spreadsheet at particular location 
                sheet.cell(row=current_row + 1, column=1).value = Name_field.get()
                sheet.cell(row=current_row + 1, column=2).value = Subject_field.get()
                sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
                sheet.cell(row=current_row + 1, column=4).value = Admn_no_field.get() 
                sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
                sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
                sheet.cell(row=current_row + 1, column=7).value = address_field.get() 
                sheet.cell(row=current_row + 1, column=8).value = Fathers_Name_field.get()
                sheet.cell(row=current_row + 1, column=9).value = Mothers_Name_field.get()
                sheet.cell(row=current_row + 1, column=10).value = Gender_field.get()
                sheet.cell(row=current_row + 1, column=11).value = Date_Of_Birth_field.get()
                sheet.cell(row=current_row + 1, column=12).value = Admn_Category_field.get()
                sheet.cell(row=current_row + 1, column=13).value = Category_field.get()
                sheet.cell(row=current_row + 1, column=14).value = Minority_field.get()
                sheet.cell(row=current_row + 1, column=15).value = Mobile_No_field.get()
                sheet.cell(row=current_row + 1, column=16).value = Email_ID_field.get()
                sheet.cell(row=current_row + 1, column=17).value = Blood_Group_field.get()
                sheet.cell(row=current_row + 1, column=18).value = Aadhar_Card_No_field.get()


                # save the file 
                wb.save('excel.xlsx') 
  
                # set focus on the name_field box 
                Name_field.focus_set() 
  
                # call the clear() function 
                #clear() 
  
  
        # Driver code 
    if __name__ == "__main__":
             # create a GUI window 
             root = Tk() 
  
             # set the background colour of GUI window 
             root.configure(background='light green') 
  
             # set the title of GUI window 
             root.title("registration form") 
                 
             # set the configuration of GUI window 
             root.geometry("470x600") 
  
             excel() 
             
             # create a Form label 
             heading = Label(root, text="Form",font='arial 24 bold', bg="light green") 
  
             # create a Name label 
             Name = Label(root, text="Name", bg="light green") 
  
             # create a Course label 
             Subject = Label(root, text="Subject", bg="light green") 
  
             # create a Semester label 
             sem = Label(root, text="Semester", bg="light green") 
  
             # create a Form No. lable 
             Admn_no = Label(root, text="Admn No.", bg="light green") 
  
             # create a Contact No. label 
             contact_no = Label(root, text="Contact No.", bg="light green") 
  
             # create a Email id label 
             email_id = Label(root, text="Email id", bg="light green") 
  
             # create a address label 
             address = Label(root, text="Address", bg="light green")

             # create a address label 
             Fathers_Name = Label(root, text="Father's Name", bg="light green")

             # create a address label 
             Mothers_Name = Label(root, text="Mother's Name", bg="light green")


             Gender = Label(root, text="Gender", bg="light green")
               
           
             Date_Of_Birth = Label(root, text="DOB", bg="light green")
           

             Admn_Category = Label(root, text="Admn Category", bg="light green")


             Category = Label(root, text="Category", bg="light green")


             Minority = Label(root, text="Minority", bg="light green")



             Mobile_No = Label(root, text="Mobile No", bg="light green")


             Email_ID = Label(root, text="Email ID", bg="light green")


             Blood_Group = Label(root, text="Blood_Group ", bg="light green")


             Aadhar_Card_No = Label(root, text="Aadhar_Card_No", bg="light green")

             # create a address label 
             Mothers_Name = Label(root, text="Mother's Name", bg="light green")


             Gender = Label(root, text="Gender", bg="light green")
                   

             Date_Of_Birth = Label(root, text="DOB", bg="light green")


             Admn_Category = Label(root, text="Admn Category", bg="light green")


             Category = Label(root, text="Category", bg="light green")


             Minority = Label(root, text="Minority", bg="light green")



             Mobile_No = Label(root, text="Mobile No", bg="light green")


             Email_ID = Label(root, text="Email ID", bg="light green")


             Blood_Group = Label(root, text="Blood_Group ", bg="light green")


             Aadhar_Card_No = Label(root, text="Aadhar_Card_No", bg="light green")


  
             # grid method is used for placing 
             # the widgets at respective positions 
             # in table like structure . 
             heading.grid(row=0, column=1) 
             Name.grid(row=1, column=0) 
             Subject.grid(row=2, column=0) 
             sem.grid(row=3, column=0) 
             Admn_no.grid(row=4, column=0) 
             contact_no.grid(row=5, column=0) 
             email_id.grid(row=6, column=0) 
             address.grid(row=7, column=0) 
             Fathers_Name.grid(row=8, column=0)
             Mothers_Name.grid(row=9, column=0)
             Gender.grid(row=10, column=0)
             Date_Of_Birth.grid(row=11, column=0)
             Admn_Category.grid(row=12, column=0)
             Category.grid(row=13, column=0)
             Minority.grid(row=14, column=0)
             Mobile_No.grid(row=15, column=0)
             Email_ID.grid(row=16, column=0)
             Blood_Group.grid(row=17, column=0)
             Aadhar_Card_No.grid(row=18, column=0)

    
             # create a text entry box 
             # for typing the information 
             Name_field = Entry(root) 
             Subject_field = Entry(root) 
             sem_field = Entry(root) 
             Admn_no_field = Entry(root) 
             contact_no_field = Entry(root) 
             email_id_field = Entry(root) 
             address_field = Entry(root) 
             Fathers_Name_field = Entry(root)
             Mothers_Name_field = Entry(root)
             Gender_field = Entry(root)
             Date_Of_Birth_field = Entry(root)
             Admn_Category_field = Entry(root)
             Category_field = Entry(root)
             Minority_field = Entry(root)
             Mobile_No_field = Entry(root)
             Email_ID_field = Entry(root)
             Blood_Group_field = Entry(root)
             Aadhar_Card_No_field = Entry(root)

    
        # bind method of widget is used for 
        # the binding the function with the events 
  
        # whenever the enter key is pressed 
        # then call the focus1 function 
             Name_field.bind("<Return>", focus1) 
  
        # whenever the enter key is pressed 
        # then call the focus2 function 
             Subject_field.bind("<Return>", focus2) 
  
        # whenever the enter key is pressed 
        # then call the focus3 function 
             sem_field.bind("<Return>", focus3) 
  
        # whenever the enter key is pressed 
        # then call the focus4 function 
             Admn_no_field.bind("<Return>", focus4) 
  
        # whenever the enter key is pressed 
        # then call the focus5 function 
             contact_no_field.bind("<Return>", focus5) 
  
        # whenever the enter key is pressed 
        # then call the focus6 function 
             email_id_field.bind("<Return>", focus6)

             Fathers_Name_field.bind("<Return>", focus7)

             Mothers_Name_field.bind("<Return>", focus9)
             Gender_field.bind("<Return>", focus10)
             Date_Of_Birth_field.bind("<Return>", focus11)
             Admn_Category_field.bind("<Return>", focus12)
             Category_field.bind("<Return>", focus13)
             Minority_field.bind("<Return>", focus14)
             Mobile_No_field.bind("<Return>", focus15)
             Email_ID_field.bind("<Return>", focus16)
             Blood_Group_field.bind("<Return>", focus17)
             Aadhar_Card_No_field.bind("<Return>", focus18)

  
             # grid method is used for placing 
        # the widgets at respective positions 
        # in table like structure . 
             Name_field.grid(row=1, column=1, ipadx="100") 
             Subject_field.grid(row=2, column=1, ipadx="100") 
             sem_field.grid(row=3, column=1, ipadx="100") 
             Admn_no_field.grid(row=4, column=1, ipadx="100") 
             contact_no_field.grid(row=5, column=1, ipadx="100") 
             email_id_field.grid(row=6, column=1, ipadx="100") 
             address_field.grid(row=7, column=1, ipadx="100") 
             Fathers_Name_field.grid(row=8, column=1, ipadx="100")
             Mothers_Name_field.grid(row=9, column=1, ipadx="100")
             Gender_field.grid(row=10, column=1, ipadx="100")
             Date_Of_Birth_field.grid(row=11, column=1, ipadx="100")
             Admn_Category_field.grid(row=12, column=1, ipadx="100")
             Category_field.grid(row=13, column=1, ipadx="100")
             Minority_field.grid(row=14, column=1, ipadx="100")
             Mobile_No_field.grid(row=15, column=1, ipadx="100")
             Email_ID_field.grid(row=16, column=1, ipadx="100")
             Blood_Group_field.grid(row=17, column=1, ipadx="100")
             Aadhar_Card_No_field.grid(row=18, column=1, ipadx="100")

             # call excel function 
             excel() 
      
             # create a Submit Button and place into the root window 
             submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
             submit.place(x=220,y=440) 
  
             # start the GUI 
             root.mainloop()
#############






























#main program started
#welcome text
top=tki.Tk()
img=Image.open("images/bg.jpg")
img=img.resize((1500,800),Image.Resampling.LANCZOS)
photo=ImageTk.PhotoImage(img)
label=tki.Label(top,image=photo)
label.grid()


top.geometry("{0}x{1}+0+0".format(top.winfo_screenwidth(), top.winfo_screenheight()))
top.title("Kendriya vidyalaya")
top.iconbitmap(r'images/Icewind Dale_1.ico')

var=StringVar()
L1=Label(top,textvariable=var,relief=RAISED,width=17,height=1,bd=0,bg="LavenderBlush",\
         activebackground="#fff",activeforeground="#42f498",fg="black")
var.set("Welcome to KV AMC")
myfont=Font(family="Impact",size=60)
L1.configure(font=myfont)
L1.place(x=330,y=20)
#end of text


#image
canvas=Canvas(top,height=140,width=150)
canvas.place(x=600,y=250)
img=PhotoImage(file="images/kvs-logo.gif")
canvas.create_image(1, 1, anchor=NW, image=img)
#image closed


#entry
password=Label(top,text="Password").place(x=570,y=505)
v1=StringVar()
E1=Entry(top,textvariable=v1,bd=5,show="*").place(x=700,y=500)
#entry closed

#message
def msg():
        messagebox.showinfo("Wait","Go to class 12")



#creating second window
def nxt(): 
    root=tki.Tk()
    root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))
    
    root.title("Kendriya vidyalaya")
    L2=Label(root,text="Student details")
    font1=Font(family="Impact",size=40)
    L2.configure(font=font1)
    L2.place(x=460,y=60)

    #menubar

    canvas=Canvas(root,height=28,width=1400,bg="#002966")
    canvas.place(x=0,y=0)


    mb=Menubutton(root,text="Menu",width=15,bg="#006600",\
                  activebackground="#006600",activeforeground="#42f498",fg="#fff")
    mb.place(x=0,y=4)
    mb.menu=Menu(mb,tearoff=0)
    mb["menu"]=mb.menu
    def logout():
            root.destroy()
            import MAINPROGRAM

    def tools():
            win=tki.Tk()
            win.geometry("{0}x{1}+0+0".format(win.winfo_screenwidth(), win.winfo_screenheight()))    
            win.title("Kendriya vidyalaya")
            L2=Label(win,text="Tools",font='arial 35 bold')            
            L2.place(x=540,y=0)
            #deifferent functions
            def graphplotter():          
                    import graphplotter
            def calculator():
                # Python program to  create a simple GUI  
                # calculator using Tkinter 
                # globally declare the expression variable 
                expression = "" 
  
  
                # Function to update expressiom 
                # in the text entry box 
                def press(num): 
                       # point out the global expression variable 
                       global expression 
  
                       # concatenation of string 
                       expression = expression + str(num) 
  
                       # update the expression by using set method 
                       equation.set(expression) 
  
  
                # Function to evaluate the final expression 
                def equalpress(): 
                        # Try and except statement is used 
                        # for handling the errors like zero 
                        # division error etc. 
  
                        # Put that code inside the try block 
                        # which may generate the error 
                        try: 
  
                            global expression 
  
                            # eval function evaluate the expression 
                            # and str function convert the result 
                            # into string 
                            total = str(eval(expression)) 
  
                            equation.set(total) 
  
                            # initialze the expression variable 
                            # by empty string 
                            expression = "" 
  
                            # if error is generate then handle 
                            # by the except block 
                        except: 
  
                                equation.set(" error ") 
                                expression = "" 
  
  
                                # Function to clear the contents 
                                # of text entry box 
                def clear(): 
                    global expression 
                    expression = "" 
                    equation.set("") 
  
  
                # Driver code 
                if __name__ == "__main__": 
                    # create a GUI window 
                    gui = Tk() 
  
                    # set the background colour of GUI window 
                    gui.configure(background="light green") 
  
                    # set the title of GUI window 
                    gui.title("Simple Calculator") 
          
                    # set the configuration of GUI window 
                    gui.geometry("400x380") 
  
                    # StringVar() is the variable class 
                    # we create an instance of this class 
                    equation = StringVar() 
  
                    # create the text entry box for 
                    # showing the expression . 
                    expression_field = Entry(gui, textvariable=equation,width=40) 
  
                    # grid method is used for placing 
                    # the widgets at respective positions 
                    # in table like structure . 
                    expression_field.place(x=60,y=50) 
  
                    equation.set('enter your expression') 
  
                    # create a Buttons and place at a particular 
                    # location inside the root window . 
                    # when user press the button, the command or 
                    # function affiliated to that button is executed . 
                    button1 = Button(gui, text=' 1 ', fg='black', bg='DodgerBlue', 
                        command=lambda: press(1), height=1, width=7) 
                    button1.place(x=30,y=100) 
  
                    button2 = Button(gui, text=' 2 ', fg='black', bg='DodgerBlue', 
                        command=lambda: press(2), height=1, width=7) 
                    button2.place(x=110,y=100) 
  
                    button3 = Button(gui, text=' 3 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(3), height=1, width=7) 
                    button3.place(x=190,y=100) 
  
                    button4 = Button(gui, text=' 4 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(4), height=1, width=7) 
                    button4.place(x=30,y=130) 
  
                    button5 = Button(gui, text=' 5 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(5), height=1, width=7) 
                    button5.place(x=110,y=130) 
  
                    button6 = Button(gui, text=' 6 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(6), height=1, width=7) 
                    button6.place(x=190,y=130) 
  
                    button7 = Button(gui, text=' 7 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(7), height=1, width=7) 
                    button7.place(x=30,y=160) 
  
                    button8 = Button(gui, text=' 8 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(8), height=1, width=7) 
                    button8.place(x=110,y=160) 
  
                    button9 = Button(gui, text=' 9 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(9), height=1, width=7) 
                    button9.place(x=190,y=160) 
  
                    button0 = Button(gui, text=' 0 ', fg='black', bg='DodgerBlue', 
                         command=lambda: press(0), height=1, width=7) 
                    button0.place(x=30,y=190)

                    button9 = Button(gui, text='Exit ', fg='black', bg='DimGrey', 
                         command=gui.destroy, height=1, width=7) 
                    button9.place(x=150,y=280)     
  
                    plus = Button(gui, text=' + ', fg='black', bg='DodgerBlue', 
                      command=lambda: press("+"), height=1, width=7) 
                    plus.place(x=270,y=100) 
  
                    minus = Button(gui, text=' - ', fg='black', bg='DodgerBlue', 
                           command=lambda: press("-"), height=1, width=7) 
                    minus.place(x=270,y=130) 
  
                    multiply = Button(gui, text=' * ', fg='black', bg='DodgerBlue', 
                          command=lambda: press("*"), height=1, width=7) 
                    multiply.place(x=270,y=160) 
  
                    divide = Button(gui, text=' / ', fg='black', bg='DodgerBlue', 
                        command=lambda: press("/"), height=1, width=7) 
                    divide.place(x=270,y=190)  
  
                    equal = Button(gui, text=' = ', fg='black', bg='DodgerBlue', 
                           command=equalpress, height=1, width=7) 
                    equal.place(x=190,y=190)  
  
                    clear = Button(gui, text='Clear', fg='black', bg='DodgerBlue', 
                           command=clear, height=1, width=7) 
                    clear.place(x=110,y=190)  
  
                    # start the GUI 
                    gui.mainloop() 


            
            button1=Button(win,text="Graph plotter",command=graphplotter,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button1.place(x=550,y=150)
            button2=Button(win,text="calculator",command=calculator,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button2.place(x=550,y=200)
            button3=Button(win,text="Exit",command=win.destroy,width=18,height=2,bd=0,bg="DimGrey",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button3.place(x=550,y=300)
            
            win.mainloop()

    mb.menu.add_checkbutton(label="Tools",command=tools)
    mb.menu.add_checkbutton(label="About")
    mb.menu.add_checkbutton(label="Logout",command=logout)    
    mb.menu.add_checkbutton(label="Exit",command=root.destroy)
    mb.place()
    

    #primary image
    canvase2=Canvas(root,height=300,width=300)
    canvase2.place(x=10,y=130)
    photo2=PhotoImage(file="images/school-kids.gif")
    canvase2.create_image(100,100,anchor=NW,image=photo2)

    #secondary image
    canvase=Canvas(root,height=300,width=300)
    canvase.place(x=900,y=130)
    photo=PhotoImage(file="images/secondary-schools.gif")
    canvase.create_image(100,100,anchor=NW,image=photo)

    #functions for buttons

    def primary(): 
        def back():
            sut.destroy()
            nxt()
        root.destroy()        
        sut=tki.Tk()
        sut.geometry("{0}x{1}+0+0".format(sut.winfo_screenwidth(), sut.winfo_screenheight()))
        sut.title("Kendriya vidyalaya")
        b=Label(sut,text='PRIMARY SECTION',font='arial 14 bold')
        myfont=Font(family="Impact",size=30)
        b.configure(font=myfont)
        b.place(x=460,y=0)
        button1=Button(sut,text="_          CLASS I         _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button1.place(x=550,y=150)
        button2=Button(sut,text="_          CLASS II        _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button2.place(x=550,y=200) 
        button3=Button(sut,text="_         CLASS III        _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button3.place(x=550,y=250)
        button4=Button(sut,text="_          CLASS IV        _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button4.place(x=550,y=300)
        button5=Button(sut,text="_          CLASS V         _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button5.place(x=550,y=350)       
        button6=Button(sut,text="Back",command=back,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button6.place(x=550,y=400)
        

        
    def secondary():
        def back1():
            nut.destroy()
            nxt()
        root.destroy()        
        nut=tki.Tk()
        nut.geometry("{0}x{1}+0+0".format(nut.winfo_screenwidth(), nut.winfo_screenheight()))
        nut.title("Kendriya vidyalaya")
        b=Label(nut,text='SECONDARY SECTION')
        myfont=Font(family="Impact",size=30)
        b.configure(font=myfont)
        b.place(x=460,y=0)
        button1=Button(nut,text="_           CLASS VI          _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button1.place(x=550,y=150)
        button2=Button(nut,text="_           CLASS VII         _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button2.place(x=550,y=200)
        button3=Button(nut,text="_           CLASS VIII        _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button3.place(x=550,y=250)
        button4=Button(nut,text="_           CLASS IX          _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button4.place(x=550,y=300) 
        button5=Button(nut,text="_           CLASS X           _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button5.place(x=550,y=350)
        button6=Button(nut,text="_           CLASS XI          _",command=msg,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button6.place(x=550,y=400)
        button7=Button(nut,text="Back",command=back1,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button7.place(x=550,y=550)

        #new window for class 12th           
        def class12():
            nut.destroy()            
            cut=tki.Tk()
            cut.geometry("{0}x{1}+0+0".format(cut.winfo_screenwidth(), cut.winfo_screenheight()))
            cut.title("Kendriya vidyalaya")
            b=Label(cut,text='Class 12th',font='arial 14 bold')
            myfont=Font(family="Impact",size=30)
            b.configure(font=myfont)
            b.place(x=525,y=0)
            #sections
####################################################################################################################################################################            
            def sciencestream():
                cut.destroy()                
                lot=tki.Tk()
                lot.geometry("{0}x{1}+0+0".format(lot.winfo_screenwidth(), lot.winfo_screenheight()))
                lot.title("Kendriya vidyalaya")
                b=Label(lot,text='Science Stream',font='arial 14 bold')
                myfont=Font(family="Impact",size=30)
                b.configure(font=myfont)
                b.place(x=500,y=0)
                #function for student data entry

                def back4():
                    lot.destroy()
                    nxt()
                    
                def studentdetails():                   
                    pii=tki.Tk()
                    pii.geometry("{0}x{1}+0+0".format(pii.winfo_screenwidth(), pii.winfo_screenheight()))
                    pii.title("Kendriya vidyalaya")
                    b=Label(pii,text='STUDENT DETAILS',font='arial 24 bold')
                    b.pack()
                    def moreinfo():
                        os.system("start EXCEL.EXE details/class12.xls")

                    def namesbio():
                        path = "details/marks.xlsx"# Give the location of the file
                        wb_obj = openpyxl.load_workbook(path)# workbook object is created 
                        sheet_obj = wb_obj.active 
                        m_row = sheet_obj.max_row
                        print("BIOLOGY STUDENTS:")# Loop will print all values of first column
                        for i in range(1, m_row + 1):
                            cell_obj = sheet_obj.cell(row = i, column =9)
                            cell_obj1 = sheet_obj.cell(row = i, column =3)
                            p=cell_obj.value
                            V=cell_obj1.value
                            if p=='NIL':
                                  print(V)
    
                    def namesmaths():
                        path = "marks.xlsx"# Give the location of the file
                        wb_obj = openpyxl.load_workbook(path)# workbook object is created 
                        sheet_obj = wb_obj.active 
                        m_row = sheet_obj.max_row
                        print("MATHEMATICS STUDENTS:")# Loop will print all values of first column
                        for i in range(1, m_row + 1):
                            cell_obj = sheet_obj.cell(row = i, column =8)
                            cell_obj1 = sheet_obj.cell(row = i, column =3)
                            p=cell_obj.value
                            V=cell_obj1.value
                            if p=='NIL':
                                  print(V)
                                  
                    button4=Button(pii,text="  BIO STUDENTS  ",command=namesbio,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button4.place(x=900,y=90)
                    button4=Button(pii,text="  MATHS STUDENTS  ",command=namesmaths,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button4.place(x=900,y=140)                    
                    button1=Button(pii,text="    More info.    ",command=moreinfo,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button1.place(x=600,y=600)
                        
                    button1=Button(pii,text="    Back    ",command=pii.destroy,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button1.place(x=600,y=650)
                    
                                              
                    Label(pii,text='CLASS: XII SCIENCE',font='arial 14 bold').place(x=10,y=60)
                    Label(pii,text='Class Teacher :  Mr. M Hamel',font='arial 12 bold').place(x=10,y=90)
                    Label(pii,text='total students :    30',font='arial 12 bold').place(x=10,y=110)
                    Label(pii,text='boys:               18',font='arial 12 bold').place(x=10,y=130)
                    Label(pii,text='girls:              12',font='arial 12 bold').place(x=10,y=150)
                
                    Label(pii,text='Subject opted',font='arial 14 bold').place(x=1100,y=60)
                    Label(pii,text='Maths + Computer science :  6',font='arial 12 bold').place(x=1100,y=90)
                    Label(pii,text='Maths + Hindi   :    4',font='arial 12 bold').place(x=1100,y=110)
                    Label(pii,text='Maths + Biology :    2',font='arial 12 bold').place(x=1100,y=130)
                    Label(pii,text='Biology + Hindi :   18',font='arial 12 bold').place(x=1100,y=150)                    
                    Label(pii,text='#The names of both bio/maths students shall be displayed on the python console.',font='arial 10 bold',bg="LavenderBlush",\
                                     activebackground="#fff",activeforeground="#42f498",fg="#ff3333").place(x=800,y=200)

                    figure2 = Figure(figsize=(4,3), dpi=100) # create a Figure 
                    subplot2 = figure2.add_subplot(111) # add a subplot
                    labels2 = 'Boys', 'Girls'
                    pieSizes = [float(18),float(12)]
                    explode2 = (0, 0.1) 
                    subplot2.pie(pieSizes, explode=explode2, labels=labels2, autopct='%1.1f%%', shadow=True, startangle=90) 
                    subplot2.axis('equal')  
                    pie2 = FigureCanvasTkAgg(figure2, pii) # create a canvas figure (matplotlib module)
                    pie2.get_tk_widget().place(x=50,y=295)
                    

                    figure = Figure(figsize=(4,3), dpi=100) # create a Figure 
                    subplot = figure.add_subplot(111) # add a subplot
                    labels = 'Maths +Cs', 'Maths + Hin ','Maths + Bio' ,'Bio + Hindi'
                    pieSizes = [float(6),float(4),float(2),float(18)]
                    explode = (0, 0,0.1,0) 
                    subplot.pie(pieSizes, explode=explode, labels=labels, autopct='%1.1f%%', shadow=True, startangle=90) 
                    subplot.axis('equal')  
                    pie = FigureCanvasTkAgg(figure, pii) # create a canvas figure (matplotlib module)
                    pie.get_tk_widget().place(x=900,y=295)
                    pii.mainloop()
                #details button closed

                def marks():                    
                    os.system("start EXCEL.EXE details/class12.xls")
          
                    
#####
                    
                    
                #function closed
                button1=Button(lot,text="    Students details    ",command=studentdetails,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button1.place(x=550,y=150)
                button2=Button(lot,text="  Student data entry  ",command=detin,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button2.place(x=550,y=200)
                button3=Button(lot,text="          Marks                 ",command=marks,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button3.place(x=550,y=250)
                button4=Button(lot,text="     Students CCA       ",width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button4.place(x=550,y=300)
                button5=Button(lot,text="          Back               ",command=back4,width=18,height=2,bd=0,bg="DimGray",\
                               activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button5.place(x=550,y=450)
##################################################################################################################################################################                              
            button1=Button(cut,text="    Science stream    ",command=sciencestream,width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button1.place(x=550,y=150)
            button2=Button(cut,text="Commerce stream  ",width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button2.place(x=550,y=200)
            button3=Button(cut,text="      Arts stream     ",width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button3.place(x=550,y=250)
            def back():
                cut.destroy()
                nxt()
                
            button4=Button(cut,text="Back",command=back,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button4.place(x=550,y=350)
            
            
        #class 12th ended    
        button7=Button(nut,text="_           CLASS XII         _",command=class12,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button7.place(x=550,y=450)


            
    #their buttons
    B1=Button(root,text="Primary classes",command=primary,width=14,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff").place(x=150,y=400)
    B2=Button(root,text="Secondary classes",command=secondary,width=14,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff").place(x=1030,y=400)

    
    B3=Button(root,text="Log out",command=logout,width=14,height=2,bd=0,bg="DimGray",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff").place(x=580,y=600)
    root.mainloop()
    
#ended
    
def insert():
    password=v1.get()
    if password=="s" :
        top.destroy()
        nxt()
    else:
        messagebox.showinfo("Login failed","wrong password")

            
B=tki.Button(top,text="Login",command=insert,width=10,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff",relief=RAISED).place(x=650,y=600)

top.mainloop()
 

