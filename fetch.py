'''
# importing openpyxl module 
import openpyxl
import xlrd
def marks2():
     path = "class12.xlsx"# Give the location of the file
     wb_obj = openpyxl.load_workbook(path) # workbook object is created 
  
     sheet_obj = wb_obj.active 
     m_row = sheet_obj.max_row 
  
    # Loop will print all values 
    # of first column  
     for i in range(1, m_row + 1):
         cell_obj = sheet_obj.cell(row = i, column = 3)
         b=cell_obj.value
         print(b)

____________________________________________________________________________    

# code for counting no. of students
# import openpyxl module 
import openpyxl 
  
# Give the location of the file 
path = "class12.xlsx"
  
# to open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
c=int(sheet_obj.max_row)-1  
# print the total number of rows 
print(c)

''' 
#successfully runned(^_^)programme to fetch names of all bio students
#similarly , for math student as well.
import openpyxl 
  
# Give the location of the file 
path = "G:\\Cs project\\shrey\\marks.xlsx"
  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row 
  
# Loop will print all values 
# of first column
print("BIOLOGY STUDENTS:")
for i in range(1, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column =9)
    cell_obj1 = sheet_obj.cell(row = i, column =3)
    p=cell_obj.value
    V=cell_obj1.value
    if p=='NIL':
        print(V)
        
#############################################################3
     
'''# Python Program - Calculate Grade of Student
print("Enter 'x' for exit.");
print("Enter marks obtained in 5 subjects: ");
mark1 = input();
if mark1 == 'x':
    exit();
else:
    mark1 = int(mark1);
    mark2 = int(input());
    mark3 = int(input());
    mark4 = int(input());
    mark5 = int(input());
    sum = mark1 + mark2 + mark3 + mark4 + mark5;
    average = sum/5;
    if(average>=91 and average<=100):
    	print("Your Grade is A+");
    elif(average>=81 and average<=90):
    	print("Your Grade is A");
    elif(average>=71 and average<=80):
    	print("Your Grade is B+");
    elif(average>=61 and average<=70):
    	print("Your Grade is B");
    elif(average>=51 and average<=60):
    	print("Your Grade is C+");
    elif(average>=41 and average<=50):
    	print("Your Grade is C");
    elif(average>=0 and average<=40):
    	print("Your Grade is F");
    else:
    	print("Strange Grade..!!");        

'''
