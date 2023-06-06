# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
  
# globally declare wb and sheet variable 
def detin():
    # opening the existing excel file 
    wb = load_workbook('excel.xlsx') 
  
    # create the sheet object 
    sheet = wb.active 
  
  
    def excel():
        # resize the width of columns in 
        # excel spreadsheet 
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10
        sheet.column_dimensions['O'].width = 10
        sheet.column_dimensions['P'].width = 30
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 20

        # write given data to an excel spreadsheet 
        # at particular location 
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "subject"
        sheet.cell(row=1, column=3).value = "Semester"
        sheet.cell(row=1, column=4).value = "Admn Number"
        sheet.cell(row=1, column=5).value = "Contact Number"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Address"
        sheet.cell(row=1, column=8).value = "Fathers Name"
        sheet.cell(row=1, column=9).value = "Mothers Name"
        sheet.cell(row=1, column=10).value = "Gender"
        sheet.cell(row=1, column=11).value = "Date Of Birth"
        sheet.cell(row=1, column=12).value = "Admn Category"
        sheet.cell(row=1, column=13).value = "Category"
        sheet.cell(row=1, column=14).value = "Minority"
        sheet.cell(row=1, column=15).value = "Mobile No"
        sheet.cell(row=1, column=16).value = "Email ID"
        sheet.cell(row=1, column=17).value = "Blood Group"
        sheet.cell(row=1, column=18).value = "Aadhar Card No"
 
    # Function to set focus (cursor) 
    def focus1(event):
            # set focus on the course_field box 
            Name_field.focus_set() 

    def focus2(event):
            # set focus on the course_field box 
            Subject_field.focus_set()
  
    # Function to set focus 
    def focus3(event): 
            # set focus on the sem_field box 
            sem_field.focus_set() 
  
  
    # Function to set focus 
    def focus4(event): 
             # set focus on the form_no_field box 
             Admn_No_field.focus_set() 
  
  
    # Function to set focus 
    def focus5(event): 
            # set focus on the contact_no_field box 
            contact_no_field.focus_set() 
  
  
    # Function to set focus 
    def focus6(event): 
            # set focus on the email_id_field box 
            email_id_field.focus_set() 
  
  
    # Function to set focus 
    def focus7(event): 
            # set focus on the address_field box 
            address_field.focus_set()

    def focus8(event): 
            # set focus on the form_no_field box 
            Fathers_Name_field.focus_set() 
    
    # Function to set focus 
    def focus9(event): 
            # set focus on the contact_no_field box 
            Mothers_Name_field.focus_set() 
  
  
    # Function to set focus 
    def focus10(event): 
            # set focus on the email_id_field box 
            Gender_field.focus_set() 
  
  
    # Function to set focus 
    def focus11(event): 
            # set focus on the address_field box 
            Date_Of_Birth_field.focus_set()

    # Function to set focus 
    def focus12(event): 
            # set focus on the address_field box 
            Admn_Category_field.focus_set()

    # Function to set focus 
    def focus13(event): 
            # set focus on the address_field box 
            Category_field.focus_set()

    # Function to set focus 
    def focus14(event): 
            # set focus on the address_field box 
            Minority_field.focus_set()

    def focus15(event): 
            # set focus on the address_field box 
            Mobile_No_field.focus_set()

    def focus16(event): 
            # set focus on the address_field box 
            Email_ID_field.focus_set()

    def focus17(event): 
            # set focus on the address_field box 
            Blood_Group_field.focus_set()

    def focus18(event): 
            # set focus on the address_field box 
            Aadhar_Card_No_field.focus_set()
 
  
  
        # Function for clearing the 
        # contents of text entry boxes 
    def clear():
             # clear the content of text entry box 
             Name_field.delete(0, END) 
             Subject_field.delete(0, END) 
             sem_field.delete(0, END) 
             Admn_No_field.delete(0, END) 
             contact_no_field.delete(0, END) 
             email_id_field.delete(0, END) 
             address_field.delete(0, END) 
             Fathers_Name_field.delete(0, END)
             Mothers_Name_field.delete(0, END)
             Gender_field.delete(0, END)
             Date_Of_Birth_field.delete(0, END)
             Admn_Category_field.delete(0, END)
             Category_field.delete(0, END)
             Minority_field.delete(0, END)
             Mobile_No_field.delete(0, END)
             Email_ID_field.delete(0, END)
             Blood_Group_field.delete(0, END)
             Aadhar_Card_No_field.delete(0, END)

    
        # Function to take data from GUI  
        # window and write to an excel file 
    def insert():
            # if user not fill any entry 
            # then print "empty input" 
            if (Name_field.get() == "" and
                Subject_field.get() == "" and
                sem_field.get() == "" and
                Admn_no_field.get() == "" and
                contact_no_field.get() == "" and
                email_id_field.get() == "" and
                address_field.get() == "" and 
                Fathers_Name_field.get() == "" and
                Mothers_Name_field.get() == "" and
                Gender_field.get() == "" and
                Date_Of_Birth_field.get() == "" and
                Admn_Category_field.get() == "" and
                Category_field.get() == "" and
                Minority_field.get() == "" and
                Mobile_No_field.get() == "" and
                Email_ID_field.get() == "" and
                Blood_Group_field.get() == "" and
                Aadhar_Card_No_field.get() == ""):

                print("empty input") 
  
            else:
                # assigning the max row and max column 
                # value upto which data is written 
                # in an excel sheet to the variable 
                current_row = sheet.max_row 
                current_column = sheet.max_column 
  
                # get method returns current text 
                # as string which we write into 
                # excel spreadsheet at particular location 
                sheet.cell(row=current_row + 1, column=1).value = Name_field.get()
                sheet.cell(row=current_row + 1, column=2).value = Subject_field.get()
                sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
                sheet.cell(row=current_row + 1, column=4).value = Admn_no_field.get() 
                sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
                sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
                sheet.cell(row=current_row + 1, column=7).value = address_field.get() 
                sheet.cell(row=current_row + 1, column=8).value = Fathers_Name_field.get()
                sheet.cell(row=current_row + 1, column=9).value = Mothers_Name_field.get()
                sheet.cell(row=current_row + 1, column=10).value = Gender_field.get()
                sheet.cell(row=current_row + 1, column=11).value = Date_Of_Birth_field.get()
                sheet.cell(row=current_row + 1, column=12).value = Admn_Category_field.get()
                sheet.cell(row=current_row + 1, column=13).value = Category_field.get()
                sheet.cell(row=current_row + 1, column=14).value = Minority_field.get()
                sheet.cell(row=current_row + 1, column=15).value = Mobile_No_field.get()
                sheet.cell(row=current_row + 1, column=16).value = Email_ID_field.get()
                sheet.cell(row=current_row + 1, column=17).value = Blood_Group_field.get()
                sheet.cell(row=current_row + 1, column=18).value = Aadhar_Card_No_field.get()


                # save the file 
                wb.save('excel.xlsx') 
  
                # set focus on the name_field box 
                Name_field.focus_set() 
  
                # call the clear() function 
                #clear() 
  
  
        # Driver code 
    if __name__ == "__main__":
             # create a GUI window 
             root = Tk() 
  
             # set the background colour of GUI window 
             root.configure(background='light green') 
  
             # set the title of GUI window 
             root.title("registration form") 
                 
             # set the configuration of GUI window 
             root.geometry("450x600") 
  
             excel() 
             
             # create a Form label 
             heading = Label(root, text="Form",font='arial 24 bold', bg="light green") 
  
             # create a Name label 
             Name = Label(root, text="Name", bg="light green") 
  
             # create a Course label 
             Subject = Label(root, text="Subject", bg="light green") 
  
             # create a Semester label 
             sem = Label(root, text="Semester", bg="light green") 
  
             # create a Form No. lable 
             Admn_no = Label(root, text="Admn No.", bg="light green") 
  
             # create a Contact No. label 
             contact_no = Label(root, text="Contact No.", bg="light green") 
  
             # create a Email id label 
             email_id = Label(root, text="Email id", bg="light green") 
  
             # create a address label 
             address = Label(root, text="Address", bg="light green")

             # create a address label 
             Fathers_Name = Label(root, text="Father's Name", bg="light green")

             # create a address label 
             Mothers_Name = Label(root, text="Mother's Name", bg="light green")


             Gender = Label(root, text="Gender", bg="light green")
               
           
             Date_Of_Birth = Label(root, text="DOB", bg="light green")
           

             Admn_Category = Label(root, text="Admn Category", bg="light green")


             Category = Label(root, text="Category", bg="light green")


             Minority = Label(root, text="Minority", bg="light green")



             Mobile_No = Label(root, text="Mobile No", bg="light green")


             Email_ID = Label(root, text="Email ID", bg="light green")


             Blood_Group = Label(root, text="Blood_Group ", bg="light green")


             Aadhar_Card_No = Label(root, text="Aadhar_Card_No", bg="light green")

             # create a address label 
             Mothers_Name = Label(root, text="Mother's Name", bg="light green")


             Gender = Label(root, text="Gender", bg="light green")
                   

             Date_Of_Birth = Label(root, text="DOB", bg="light green")


             Admn_Category = Label(root, text="Admn Category", bg="light green")


             Category = Label(root, text="Category", bg="light green")


             Minority = Label(root, text="Minority", bg="light green")



             Mobile_No = Label(root, text="Mobile No", bg="light green")


             Email_ID = Label(root, text="Email ID", bg="light green")


             Blood_Group = Label(root, text="Blood_Group ", bg="light green")


             Aadhar_Card_No = Label(root, text="Aadhar_Card_No", bg="light green")


  
             # grid method is used for placing 
             # the widgets at respective positions 
             # in table like structure . 
             heading.grid(row=0, column=1) 
             Name.grid(row=1, column=0) 
             Subject.grid(row=2, column=0) 
             sem.grid(row=3, column=0) 
             Admn_no.grid(row=4, column=0) 
             contact_no.grid(row=5, column=0) 
             email_id.grid(row=6, column=0) 
             address.grid(row=7, column=0) 
             Fathers_Name.grid(row=8, column=0)
             Mothers_Name.grid(row=9, column=0)
             Gender.grid(row=10, column=0)
             Date_Of_Birth.grid(row=11, column=0)
             Admn_Category.grid(row=12, column=0)
             Category.grid(row=13, column=0)
             Minority.grid(row=14, column=0)
             Mobile_No.grid(row=15, column=0)
             Email_ID.grid(row=16, column=0)
             Blood_Group.grid(row=17, column=0)
             Aadhar_Card_No.grid(row=18, column=0)

    
             # create a text entry box 
             # for typing the information 
             Name_field = Entry(root) 
             Subject_field = Entry(root) 
             sem_field = Entry(root) 
             Admn_no_field = Entry(root) 
             contact_no_field = Entry(root) 
             email_id_field = Entry(root) 
             address_field = Entry(root) 
             Fathers_Name_field = Entry(root)
             Mothers_Name_field = Entry(root)
             Gender_field = Entry(root)
             Date_Of_Birth_field = Entry(root)
             Admn_Category_field = Entry(root)
             Category_field = Entry(root)
             Minority_field = Entry(root)
             Mobile_No_field = Entry(root)
             Email_ID_field = Entry(root)
             Blood_Group_field = Entry(root)
             Aadhar_Card_No_field = Entry(root)

    
        # bind method of widget is used for 
        # the binding the function with the events 
  
        # whenever the enter key is pressed 
        # then call the focus1 function 
             Name_field.bind("<Return>", focus1) 
  
        # whenever the enter key is pressed 
        # then call the focus2 function 
             Subject_field.bind("<Return>", focus2) 
  
        # whenever the enter key is pressed 
        # then call the focus3 function 
             sem_field.bind("<Return>", focus3) 
  
        # whenever the enter key is pressed 
        # then call the focus4 function 
             Admn_no_field.bind("<Return>", focus4) 
  
        # whenever the enter key is pressed 
        # then call the focus5 function 
             contact_no_field.bind("<Return>", focus5) 
  
        # whenever the enter key is pressed 
        # then call the focus6 function 
             email_id_field.bind("<Return>", focus6)

             Fathers_Name_field.bind("<Return>", focus7)

             Mothers_Name_field.bind("<Return>", focus9)
             Gender_field.bind("<Return>", focus10)
             Date_Of_Birth_field.bind("<Return>", focus11)
             Admn_Category_field.bind("<Return>", focus12)
             Category_field.bind("<Return>", focus13)
             Minority_field.bind("<Return>", focus14)
             Mobile_No_field.bind("<Return>", focus15)
             Email_ID_field.bind("<Return>", focus16)
             Blood_Group_field.bind("<Return>", focus17)
             Aadhar_Card_No_field.bind("<Return>", focus18)

  
             # grid method is used for placing 
        # the widgets at respective positions 
        # in table like structure . 
             Name_field.grid(row=1, column=1, ipadx="100") 
             Subject_field.grid(row=2, column=1, ipadx="100") 
             sem_field.grid(row=3, column=1, ipadx="100") 
             Admn_no_field.grid(row=4, column=1, ipadx="100") 
             contact_no_field.grid(row=5, column=1, ipadx="100") 
             email_id_field.grid(row=6, column=1, ipadx="100") 
             address_field.grid(row=7, column=1, ipadx="100") 
             Fathers_Name_field.grid(row=8, column=1, ipadx="100")
             Mothers_Name_field.grid(row=9, column=1, ipadx="100")
             Gender_field.grid(row=10, column=1, ipadx="100")
             Date_Of_Birth_field.grid(row=11, column=1, ipadx="100")
             Admn_Category_field.grid(row=12, column=1, ipadx="100")
             Category_field.grid(row=13, column=1, ipadx="100")
             Minority_field.grid(row=14, column=1, ipadx="100")
             Mobile_No_field.grid(row=15, column=1, ipadx="100")
             Email_ID_field.grid(row=16, column=1, ipadx="100")
             Blood_Group_field.grid(row=17, column=1, ipadx="100")
             Aadhar_Card_No_field.grid(row=18, column=1, ipadx="100")

             # call excel function 
             excel() 
      
             # create a Submit Button and place into the root window 
             submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
             submit.place(x=220,y=440) 
  
             # start the GUI 
             root.mainloop()

detin()
